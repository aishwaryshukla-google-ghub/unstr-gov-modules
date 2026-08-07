import io
from .base import BaseDataHandler


class ExcelHandler(BaseDataHandler):
    """
    Handler for Excel spreadsheets (.xlsx, .xls).
    Iterates through worksheets and converts tabular data into styled PDF pages.
    """

    def convert_to_pdf(self, file_bytes: bytes, filename: str) -> bytes:
        try:
            import openpyxl
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=landscape(letter),
                rightMargin=20,
                leftMargin=20,
                topMargin=20,
                bottomMargin=20
            )

            styles = getSampleStyleSheet()
            sheet_title_style = ParagraphStyle(
                "SheetTitle",
                parent=styles["Heading2"],
                fontSize=13,
                leading=15,
                textColor=colors.HexColor("#1a73e8"),
                spaceBefore=10,
                spaceAfter=6
            )
            cell_style = ParagraphStyle(
                "ExcelCell",
                parent=styles["Normal"],
                fontSize=7,
                leading=9,
                textColor=colors.HexColor("#202124")
            )
            header_style = ParagraphStyle(
                "ExcelHeader",
                parent=styles["Normal"],
                fontSize=7,
                leading=9,
                fontName="Helvetica-Bold",
                textColor=colors.white
            )

            story = []

            for sheet_name in wb.sheetnames[:5]:  # Limit to first 5 sheets
                ws = wb[sheet_name]
                story.append(Paragraph(f"Sheet: {sheet_name} ({filename})", sheet_title_style))

                table_data = []
                for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
                    if r_idx > 150:  # Cap at 150 rows per sheet
                        break
                    if not any(row):  # Skip completely empty rows
                        continue

                    formatted_row = []
                    for cell_val in row[:15]:  # Cap at 15 columns
                        val_str = str(cell_val) if cell_val is not None else ""
                        safe_val = val_str.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                        p = Paragraph(safe_val, header_style if r_idx == 0 else cell_style)
                        formatted_row.append(p)
                    table_data.append(formatted_row)

                if table_data:
                    t = Table(table_data, repeatRows=1)
                    t.setStyle(TableStyle([
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a73e8")),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dadce0")),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
                    ]))
                    story.append(t)
                    story.append(Spacer(1, 12))

            doc.build(story)
            return buffer.getvalue()

        except ImportError:
            # Fallback simple text-based rendering
            from .text_md_handler import TextMarkdownHandler
            return TextMarkdownHandler().convert_to_pdf(b"Excel file preview (openpyxl not installed).", filename)
